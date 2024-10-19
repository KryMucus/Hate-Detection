from openpyxl import load_workbook
from nltk.stem.porter import *
import pandas as pd
import html
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
import numpy as np
import argparse
import os
import warnings
from termcolor import colored

warnings.filterwarnings("ignore", category=pd.errors.SettingWithCopyWarning)

def Create_Datasets():

    # Load the workbook and select the active worksheet
    for i in range (1,9):
        wb = load_workbook(f'Group{i}.xlsx')
        ws = wb.active

        # Iterate over cells in the 'Content' column (assuming it's the first column)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if isinstance(cell.value, str):
                cell.value = html.unescape(cell.value)

        # Save changes back to the file
        wb.save(f'Group{i}_modified.xlsx')

    # Group 1
    df_1 = pd.read_excel("Group1_modified.xlsx")
    df_1.dataframeName = 'Group 1'
    #Name these columns
    df_1.columns=["Content","","","","","","","","","","","","","","","","","","","","","Label","","","",""]
    #Keep only those
    df_1 = df_1[['Content','Label']]
    df_1 = df_1.drop(index=[0,1])

    # Group 2
    df_2 = pd.read_excel("Group2_modified.xlsx")
    df_2.dataframeName = 'Group 2'
    df_2.columns= ["Content","Label"]
    
    # Group 3
    df_3 = pd.read_excel("Group3_modified.xlsx")
    df_3.dataframeName = 'Group 3'
    df_3.columns=['Content',"Brian",'Georgia','Label','Agreement']
    df_3 = df_3[['Content','Label']]

    # Group 4
    df_4 = pd.read_excel("Group4_modified.xlsx")
    df_4.dataframeName = 'Group 4'
    df_4.columns = ['Content','Label']
    df_4.head()

    # Group 5
    df_5 = pd.read_excel("Group5_modified.xlsx")
    df_5.dataframeName = 'Group 5'
    df_5.columns = ['Content','Label']
    df_5.drop(index=0)

    # Group 6
    df_6 = pd.read_excel("Group6_modified.xlsx", sheet_name="FINAL Labelled Dataset")
    df_6.dataframeName = 'Group 6'
    df_6.columns = ['Label','Content']
    # Switch columns values
    m = df_6['Label'] != "NX"
    df_6.loc[m, ['Label','Content']] = df_6.loc[m, ['Content','Label']].values
    # Switch columns names
    df_6.columns= ["Content","Label"]

    # Group 7 
    df_7 = pd.read_excel("Group7_modified.xlsx")
    df_7.dataframeName = 'Group 7'
    df_7.columns = ['Content','Label']
    df_7.drop(index=0)

    # Group 8
    df_8 = pd.read_excel("Group8_modified.xlsx")
    df_8.dataframeName = 'Group 8'
    df_8.columns = ['Content','Label']

    # Group 9
    df_9 = pd.read_excel("Group9_modified.xlsx")
    df_9.dataframeName = 'Group 9'
    df_9.columns=['Number',"Content",'Label','IRR']
    df_9 = df_9[['Content','Label']]

    merged_df = pd.concat([df_1, df_2, df_3, df_4, df_5, df_6,df_7, df_8, df_9], ignore_index=True)
    return merged_df

def binarise(label):
    if label.replace(" ", "") == "fake" :
        return 1
    elif (label.replace(" ", "") == "notfake") or label.replace(" ", "") == "not-fake" :
        return 0
    else :
        return 2
    
def create_bow_train(tweet):
    stopwords = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'theirs', 'themselves', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those', 'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'having', 'do', 'does', 'did', 'doing', 'a', 'an', 'the', 'and', 'but', 'if', 'or', 'because', 'as', 'until', 'while', 'of', 'at', 'by', 'for', 'with', 'about', 'against', 'between', 'into', 'through', 'during', 'before', 'after', 'above', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'all', 'any', 'both', 'each', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very', 's', 't', 'can', 'will', 'just', 'don', 'should', 'now',"#ff", "ff", "rt"]
    # Removal of extra spaces
    tweet = re.sub(r'\s+', ' ', tweet)
    # Remove hashtags
    #tweet = re.sub(r'#\w+', ' ', tweet)
    # Removal of @name [mention]
    tweet = re.sub(r'@[\w\-]+', ' ', tweet)
    # Removal of links [https://abc.com]
    tweet = re.sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', ' ', tweet)
    # Removal of punctuations and numbers
    tweet = re.sub(r'[^a-zA-Z]', ' ', tweet)
    # Remove whitespace with a single space
    tweet = re.sub(r'\s+', ' ', tweet)
    # Remove leading and trailing whitespace
    tweet = tweet.strip()
    # Removal of capitalization
    tweet = tweet.lower()
    # Tokenizing
    tokenized_tweet = tweet.split()
    tweet_bow=[]
    # Stopwords removal
    for word in tokenized_tweet:
        if word not in stopwords :
            #No stemming involved in BOW
            #stemmed_word = stemmer.stem(word)
            #processed_tweet.append(stemmed_word)
            tweet_bow.append(word)
            if word not in total_bow :
                total_bow.append(word)
    # Join the processed words back into a single string
    # return ' '.join(processed_tweet)
    return tweet_bow

def create_bow_test(tweet):
    stopwords = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'theirs', 'themselves', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those', 'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'having', 'do', 'does', 'did', 'doing', 'a', 'an', 'the', 'and', 'but', 'if', 'or', 'because', 'as', 'until', 'while', 'of', 'at', 'by', 'for', 'with', 'about', 'against', 'between', 'into', 'through', 'during', 'before', 'after', 'above', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'all', 'any', 'both', 'each', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very', 's', 't', 'can', 'will', 'just', 'don', 'should', 'now',"#ff", "ff", "rt"]
    # Removal of extra spaces
    tweet = re.sub(r'\s+', ' ', tweet)
    # Remove hashtags
    #tweet = re.sub(r'#\w+', ' ', tweet)
    # Removal of @name [mention]
    tweet = re.sub(r'@[\w\-]+', ' ', tweet)
    # Removal of links [https://abc.com]
    tweet = re.sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', ' ', tweet)
    # Removal of punctuations and numbers
    tweet = re.sub(r'[^a-zA-Z]', ' ', tweet)
    # Remove whitespace with a single space
    tweet = re.sub(r'\s+', ' ', tweet)
    # Remove leading and trailing whitespace
    tweet = tweet.strip()
    # Removal of capitalization
    tweet = tweet.lower()
    # Tokenizing
    tokenized_tweet = tweet.split()
    # Stopwords removal
    tweet_bow=[]
    for word in tokenized_tweet:
        if word not in stopwords and word in total_bow :
            tweet_bow.append(word)
    return tweet_bow

def vectorize(BoW):
    vector=[]
    for w in total_bow:
        vector.append(BoW.count(w))
    #print ("Vector type : ", type(vector))
    return vector
    
def preprocessing(merged_df):
    global total_bow
    total_bow = []
    # Lowering the characters 
    merged_df= merged_df[merged_df['Label'].apply(lambda x: isinstance(x, (str, bytes)))]
    merged_df['Label'] = merged_df['Label'].apply(str.lower)
    # Convert the Label to 0 (Not Fake) and 1 (Fake)
    merged_df['Binarised Label'] = merged_df['Label'].apply(binarise)
    merged_df['Binarised Label'].value_counts()
    # Make sure all the content column is made of strings
    merged_df["Content"] = merged_df["Content"].apply(str)
    merged_df['BoW'] = merged_df["Content"].apply(create_bow_train)
    # Let's drop all the entries that aren't unique, as they would mess with our model  
    merged_df = merged_df.drop_duplicates(subset='BoW')
    merged_df["Vector"] = merged_df["BoW"].apply(vectorize)
    return merged_df


def is_valid_file(parser, arg):
    if not os.path.exists(arg):
        parser.error("The file %s does not exist!" % arg)
    else:
        print("File found !")
        return open(arg, 'r')  # return an open file handle
    
def Predict(test_list, bowed_df):

    vectorized_tweets = []
    for tweet in test_list:
        # No empty tweets
        if tweet.replace(" ","").replace("\n","") != "" :
            processed_tweet = create_bow_test(tweet)
            vector = vectorize(processed_tweet)
            vectorized_tweets.append(vector)
    test_vectors = np.array(vectorized_tweets)
    X = np.array(bowed_df["Vector"].tolist()) 
    y = bowed_df["Binarised Label"]           
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    mnb = MultinomialNB()
    mnb.fit(X_train, y_train)
    y_pred = mnb.predict(test_vectors)
    for idx, label in enumerate(y_pred):
        print("\n")
        print("According to my expertise, the following tweet :", colored(test_list[idx], "light_cyan"))
        if label == 1:
            print(colored("is FAKE.", 'red'))
        else:
            print(colored("is NOT FAKE.", 'green'))

parser = argparse.ArgumentParser(
    prog="predict.py",
    description="Written by Jules Deleuse",
    usage= "python3 predict.py  <Argument File>",
    epilog="Thanks for using this program")
parser.add_argument('Filename', type=lambda x: is_valid_file(parser, x))
args = parser.parse_args()
file = args.Filename.read()
test_list = file.split("\n")
print("Merging Datasets...")
merged_df = Create_Datasets()
print("Creating the BoW...")
bowed_df = preprocessing(merged_df)
print("Algorithm ready...")
print("Predicting ...")
results = Predict (test_list, bowed_df)
print (results) 


